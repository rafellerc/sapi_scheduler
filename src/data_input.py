from collections import namedtuple
import datetime

import numpy as np
from openpyxl import load_workbook


class ConsistencyBetweenFilesError(Exception):
    pass


Person = namedtuple(
    'Person', ['name', 'gender', 'exp_level', 'task_answ'])

ProblemInfo = namedtuple('ProblemInfo', [
                         'n_p_units', 'n_days', 'n_tasks', 'tasks', 'ids',
                         'names', 'days', 'people_per_task', 'women_per_task',
                         'parents_per_task', 'n_people', 'n_women', 'n_parents',
                         'reject', 'force'])

FichaServoInfo = namedtuple('FichaServoInfo', [
                            'id_column', 'exp_column', 'gend_column',
                            'task_columns', 'name_column'])

DemandaInfo = namedtuple('DemandaInfo', [
                         'people_column', 'women_column',
                         'parents_column'])

IndispInfo = namedtuple('IndispInfo', ['id_column', 'name_column',
                                       'day_columns'])


DEFAULT_FICHA_SERVO_INFO = FichaServoInfo(id_column='A',
                                          exp_column='B',
                                          gend_column='S',
                                          task_columns=['G', 'H', 'I', 'J', 'K',
                                                        'L', 'M', 'N', 'O'],
                                          name_column='R')

DEFAULT_DEMANDA_INFO = DemandaInfo(people_column='B', women_column='C',
                                   parents_column='D')

DEFAULT_INDISP_INFO = IndispInfo(id_column='A', name_column='C',
                                 day_columns=['D', 'E', 'F', 'G'])

# TODO Add some checks


def read_ficha_servo(file_path, id_column='A', exp_column='B', gend_column='S',
                     task_columns=['G', 'H', 'I', 'J',
                                   'K', 'L', 'M', 'N', 'O'],
                     name_column='R'):
    """ Reads the Ficha Servo spreadsheet.
    Args:
        file_path: (str) The full path to the 'ficha_servo' file.
        id_column: (str) The column containing the id for the people.
        exp_column: (str) The column containing the experience level.
        gend_column: (str) The column containing the gender of the person.
        name_column: (str) The column containing the name of each person.
        task_columns: (list) A list containing all the letters of the columns
    containing tasks.
    Returns:
        tasks: (list) The list with all the available tasks.
        people: (dict) The dict of form {id:PersonNamedTuple}
    """
    wb = load_workbook(file_path)
    sheet = wb.active

    column_to_task = {}
    for column in task_columns:
        column_to_task[column] = sheet[column + '1'].value
    tasks = list(column_to_task.values())

    idx = 2
    people = {}
    # NOTE Here I use a max number of lines equal to 1000
    while sheet[id_column + str(idx)].value is not None and idx < 1000:
        task_answ = {
            column_to_task[column]: sheet[column + str(idx)].value for column in task_columns}
        person = Person(name=sheet[name_column + str(idx)].value,
                        gender=sheet[gend_column + str(idx)].value,
                        exp_level=sheet[exp_column + str(idx)].value,
                        task_answ=task_answ)
        people[sheet[id_column + str(idx)].value] = person
        idx += 1
    return tasks, people


def read_demanda(file_path, people_column='B', women_column='C', parents_column='D'):
    """ Reads the demanda file that specifies the number of people demanded
    by each task, as well as the number of women and experienced people.
    Args:
    Returns:

    """
    wb = load_workbook(file_path)
    sheet = wb.active
    idx = 2
    tasks = []
    people_demand = {}
    women_demand = {}
    parents_demand = {}
    while sheet['A' + str(idx)].value is not None and idx < 1000:
        tasks.append(sheet['A' + str(idx)].value)
        people_demand[sheet['A' +
                            str(idx)].value] = sheet[people_column + str(idx)].value
        women_demand[sheet['A' +
                           str(idx)].value] = sheet[women_column + str(idx)].value
        parents_demand[sheet['A' +
                             str(idx)].value] = sheet[parents_column + str(idx)].value
        idx += 1
    return tasks, people_demand, women_demand, parents_demand


def read_solution_sheet(file_path, id_column='A', name_column='C',
                        day_columns=['D', 'E', 'F', 'G']):
    """ Reads the solution sheet with the indisponibilities and pre-allocations,
    as well as previous days infos to define completely the instance.
    Args:
        file_path: (str) The full path to the solution sheet.
        id_column: (str) The column containing the id of the people
        name_column: (str) The column containing the names of each person.
        day_columns: (list) A list containing the columns of the days to be solved.
    Returns:
        days: (list) A list with the names of the days
        sheet_info: (list) A list with the triples (id, day name, info) where
            info can be one of the tasks or an indisponibility.
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    col_to_day = {}
    for column in day_columns:
        value = sheet[column + '1'].value
        if isinstance(value, datetime.datetime):
            col_to_day[column] = value.strftime('%d/%m/%Y')
        elif value is str:
            col_to_day = value
        else:
            raise Exception('Unknown type for date')

    sheet_info = []
    idx = 2
    while sheet[id_column + str(idx)].value is not None and idx < 1000:
        id_ = sheet[id_column + str(idx)].value
        for column in day_columns:
            day = col_to_day[column]
            info = sheet[column + str(idx)].value
            if info is not None:
                sheet_info.append((id_, day, info))
        idx += 1
    days = list(col_to_day.values())
    return days, sheet_info


def get_problem_info(ficha_servo_path, demanda_path, sol_sheet_path,
                     exp_threshold=3, ficha_servo_info=DEFAULT_FICHA_SERVO_INFO,
                     demanda_info=DEFAULT_DEMANDA_INFO, indip_info=DEFAULT_INDISP_INFO):
                    #  solution_sheet_day_columns=['D', 'E', 'F', 'G']):
    """ Gets the complete info from the sheets to define the problem
    Args:
        ficha_servo_path: (str) The path to the ficha servo excel file.
        demanda_path: (str) The path to the demanda excel file.
        sol_sheet_path: (str) The path to the indisponibilidades excel file.
        exp_threshold: (str) The number (between 0 and 5) corresponding to
            the lowest experience level that defines someone as experienced or
            parent.
    Returns:
        problem_info: (namedtuple) The named tuple with the problem information.
    """
    tasks, people = read_ficha_servo(ficha_servo_path, ficha_servo_info.id_column,
                                     ficha_servo_info.exp_column, ficha_servo_info.gend_column,
                                     ficha_servo_info.task_columns, ficha_servo_info.name_column)
    tasks_check, people_demand, women_demand, parents_demand = read_demanda(
        demanda_path, demanda_info.people_column, demanda_info.women_column,
        demanda_info.parents_column)
    days, sheet_info = read_solution_sheet(
        sol_sheet_path, indip_info.id_column, indip_info.name_column, indip_info.day_columns)

    if tasks != tasks_check:
        raise ConsistencyBetweenFilesError("The names of the tasks are different "
                                           "between the 'ficha servo' and 'demanda' files")
    n_p_units = len(people)
    n_days = len(days)
    n_tasks = len(tasks)

    # The list is extracted from the dict following the 'tasks' list, in order
    # to guarantee they are both in the same order.
    people_demand_list = [people_demand[task] for task in tasks]
    women_demand_list = [women_demand[task] for task in tasks]
    parents_demand_list = [parents_demand[task] for task in tasks]

    # NOTE The repetition assumes that the demand is a single list, i.e., that
    # the demands are specified only once, and equal on all days.
    people_per_task = np.repeat(np.array(people_demand_list)[
        np.newaxis, :], n_days, axis=0)
    women_per_task = np.repeat(np.array(women_demand_list)[
        np.newaxis, :], n_days, axis=0)
    parents_per_task = np.repeat(np.array(parents_demand_list)[
        np.newaxis, :], n_days, axis=0)

    id_to_i = {}
    names = []
    n_people = []
    n_women = []
    n_parents = []
    reject = []
    force = []
    for i, (id_, person) in enumerate(people.items()):
        id_to_i[id_] = i
        names.append(person.name)
        # NOTE since I'm not yet allowing a multi-person p_unit, every p_unit
        # has 1 and only 1 person on it.
        n_people.append(1)
        n_women.append(1) if person.gender == 'F' else n_women.append(0)
        n_parents.append(1) if person.exp_level >= exp_threshold \
            else n_parents.append(0)
        for task, answ in person.task_answ.items():
            k = tasks.index(task)
            if answ == "Não Aceito":
                # Add a rejection for the given task on each day
                # NOTE: So far I'm not using the distiction between 'Gosto' and
                # 'Aceito'.
                reject += [(i, j, k) for j in range(len(days))]
    n_people = np.array(n_people)
    n_women = np.array(n_women)
    n_parents = np.array(n_parents)
    for id_, day, info in sheet_info:
        i = id_to_i[id_]
        j = days.index(day)
        # If the cell contains a task, then the person is forced to perform it
        # on the given day
        if info in tasks:
            k = tasks.index(info)
            force.append((i, j, k))
        # If the cell contains 'indisp' the person is unavailable during that
        # day and thus rejected on all tasks for that day.
        elif info == 'indisp':
            reject += [(i, j, k) for k in range(len(tasks))]

    ids = [None] * len(id_to_i)
    for key, value in id_to_i.items():
        ids[value] = key

    # Remove repeated triples.
    force = list(set(force))
    reject = list(set(reject))

    problem_info = ProblemInfo(n_p_units=n_p_units,
                               n_days=n_days,
                               n_tasks=n_tasks,
                               tasks=tasks,
                               ids=ids,
                               names=names,
                               days=days,
                               people_per_task=people_per_task,
                               women_per_task=women_per_task,
                               parents_per_task=parents_per_task,
                               n_people=n_people,
                               n_women=n_women,
                               n_parents=n_parents,
                               reject=reject,
                               force=force)

    return problem_info


if __name__ == '__main__':
    PATH = 'C:\\Users\\Rafael\\dev\\sapi_scheduler\\data\\Example_Experiencia.xlsx'
    tasks, people = read_ficha_servo(PATH)
    print(tasks)
