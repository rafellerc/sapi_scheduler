from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np
import yaml
from ast import literal_eval
import os.path as osp
from os import mkdir


from src.instance import Instance
from src.data_input import get_problem_info, ProblemInfo, FichaServoInfo, DemandaInfo, IndispInfo


def sol_to_array(solution, prb_info):
    """ Transforms a solution into an array of values ready to be written to
    an excel file. Basically going day by day and task by task checking who
    is allocated and writing all the names on each cell.

    Args:
        solution: (numpy.ndarray) The n_p_units*n_days*n_tasks solution matrix.
        prb_info: (namedtuple) The problem information named tuple defined in
            data_input.py
    Returns:
        sol_array: (list) The array (list of lists) containing the values for
            each cell in the solution excel sheet.
        days_worked: (list) The list with the number of days each person worked.
            The elements are ordered in the same order as the 'names' in the
            prb_info named tuple.
    """
    sol_array = []
    headers = [' '] + prb_info.days

    days_worked = [0] * len(prb_info.names)
    sol_array.append(headers)
    for k in range(solution.shape[2]):
        row = [prb_info.tasks[k]]
        for j in range(solution.shape[1]):
            people = ''
            for i in range(solution.shape[0]):
                if solution[i, j, k] == 1:
                    people = people + str(prb_info.names[i]) + ', '
                    days_worked[i] += 1
            if people.endswith(', '):
                people = people[0:-2]
            row.append(people)
        sol_array.append(row)
    return sol_array, days_worked


def make_group(list_of_ids, prb_info):
    """ Takes a problem_info tuple (defined in data_input.py) and a
    pair (or any number) of ids to make a new prb_info with this pair as a
    couple/group (several people on the same p_unit).

    Args:
        list_of_ids: (list) A list of ids, representing
            a new couple/group to be created.
        prb_info: (namedtuple) The current problem info named tuple.
    Returns:
        new_prb_info: (namedtuple) The new problem info with the new
            couple/group and all the necessary changes (removing the people
            as individuals, changing the 'names' and 'id' fields, etc)
    """
    names = prb_info.names[:]
    instance_ids = prb_info.ids[:]
    force = prb_info.force[:]
    reject = prb_info.reject[:]

    people = 0
    women = 0
    parents = 0
    new_names = ''
    new_ids = []
    n_p_units = prb_info.n_p_units

    n_people = list(np.copy(prb_info.n_people))
    n_women = list(np.copy(prb_info.n_women))
    n_parents = list(np.copy(prb_info.n_parents))

    first_name = True

    group_index = n_p_units - len(list_of_ids)
    indexes_of_group = []

    # This list is used to capture the change in the overal positions of the
    # people. To find the new index based on an old index just call:
    # index_struct.index(old_index)
    index_struct = list(range(n_p_units))

    for id_ in list_of_ids:
        index = instance_ids.index(id_)
        indexes_of_group.append(index_struct.pop(index))
        people += n_people.pop(index)
        women += n_women.pop(index)
        parents += n_parents.pop(index)
        if not first_name:
            new_names += ', '
        first_name = False
        new_names += names.pop(index)
        new_ids.append(instance_ids.pop(index))
        n_p_units -= 1

    for i, triple in enumerate(prb_info.force):
        if triple[0] in indexes_of_group:
            force[i] = (group_index, triple[1], triple[2])
        else:
            force[i] = (index_struct.index(triple[0]), triple[1], triple[2])
    for i, triple in enumerate(prb_info.reject):
        if triple[0] in indexes_of_group:
            reject[i] = (group_index, triple[1], triple[2])
        else:
            reject[i] = (index_struct.index(triple[0]), triple[1], triple[2])

    force = list(set(force))
    reject = list(set(reject))
    force = [element for element in force if element not in reject]
    n_people = np.array(n_people + [people])
    n_women = np.array(n_women + [women])
    n_parents = np.array(n_parents + [parents])
    instance_ids.append(new_ids)
    names.append(new_names)
    n_p_units += 1

    new_prb_info = ProblemInfo(n_p_units=n_p_units,
                               n_days=prb_info.n_days,
                               n_tasks=prb_info.n_tasks,
                               tasks=prb_info.tasks,
                               ids=instance_ids,
                               names=names,
                               days=prb_info.days[:],
                               people_per_task=np.copy(
                                   prb_info.people_per_task),
                               women_per_task=np.copy(prb_info.women_per_task),
                               parents_per_task=np.copy(
                                   prb_info.parents_per_task),
                               n_people=n_people,
                               n_women=n_women,
                               n_parents=n_parents,
                               reject=reject,
                               force=force)

    return new_prb_info


def main():
    # Open the configuration file
    with open('config.yml', 'r') as ymlfile:
        cfg = yaml.load(ymlfile, Loader=yaml.SafeLoader)

    ficha_path = cfg['ficha_servo_path']
    demanda_path = cfg['demanda_path']
    indisp_path = cfg['indisp_path']

    ficha_info = FichaServoInfo(id_column=cfg['ficha_servo_info']['id_column'],
                                exp_column=cfg['ficha_servo_info']['exp_column'],
                                gend_column=cfg['ficha_servo_info']['gend_column'],
                                task_columns=cfg['ficha_servo_info']['task_columns'],
                                name_column=cfg['ficha_servo_info']['name_column'])

    demanda_info = DemandaInfo(people_column=cfg['demanda_info']['people_column'],
                               women_column=cfg['demanda_info']['women_column'],
                               parents_column=cfg['demanda_info']['parents_column'])

    indisp_info = IndispInfo(id_column=cfg['indisp_info']['id_column'],
                             name_column=cfg['indisp_info']['name_column'],
                             day_columns=cfg['indisp_info']['day_columns'])

    # Get the problem info from the excel files
    prb_info = get_problem_info(ficha_path, demanda_path, indisp_path,
                                exp_threshold=cfg['exp_threshold'], ficha_servo_info=ficha_info,
                                demanda_info=demanda_info, indip_info=indisp_info)

    # Add the couples/groups to the problem info
    couples = [literal_eval(couple) for couple in cfg['couples']]

    for couple in couples:
        prb_info = make_group(couple, prb_info)

    # Make the problem instance from the problem info
    instance = Instance(prb_info.n_p_units,
                        prb_info.n_days, prb_info.n_tasks,
                        max_sols=cfg['maximum_number_of_solutions'])
    instance.n_people = prb_info.n_people
    instance.n_women = prb_info.n_women
    instance.n_parents = prb_info.n_parents
    instance.people_per_task = prb_info.people_per_task
    instance.women_per_task = prb_info.women_per_task
    instance.parents_per_task = prb_info.parents_per_task
    instance.force = prb_info.force
    instance.reject = prb_info.reject

    # Solve the instance
    instance.solve(one_alloc_period=cfg['one_allocation_every_how_many_weeks'])

    # Write the solutions to excel file(s)
    solutions_path = osp.join(osp.dirname(osp.abspath(__file__)), 'solutions')

    single_file = cfg['solutions_in_single_file']
    if single_file:
        wb = Workbook()
        dest_filename = osp.join(
            solutions_path, '{}.xlsx'.format(cfg['solution_name']))
        remove_sheet = wb.active
        remove_sheet.title = 'remove'
    else:
        solution_dir = dest_filename = osp.join(
            solutions_path, cfg['solution_name'])
        mkdir(solution_dir)

    for i, solution in enumerate(instance.solution_list):
        solution_array, days_worked = sol_to_array(solution, prb_info)

        if single_file:
            sheet = wb.create_sheet('Solution{}'.format(i))
        else:
            wb = Workbook()
            dest_filename = osp.join(
                solution_dir, 'solution_{}.xlsx'.format(i))
            sheet = wb.active
            sheet.title = "Solution{}".format(i)

        for row in range(len(solution_array)):
            for column in range(len(solution_array[0])):
                sheet.cell(column=column + 1, row=row + 1,
                           value=solution_array[row][column])

        p_counter_column = len(solution_array[0]) + 2
        sheet.cell(column=p_counter_column, row=1,
                   value='Name')
        sheet.cell(column=p_counter_column + 1, row=1,
                   value='Number of days worked')
        for j, name in enumerate(prb_info.names):
            sheet.cell(column=p_counter_column, row=j + 2,
                       value=name)
            sheet.cell(column=p_counter_column + 1, row=j + 2,
                       value=days_worked[j])

        if not single_file:
            wb.save(filename=dest_filename)

    if single_file:
        wb.remove(wb.get_sheet_by_name('remove'))
        wb.save(filename=dest_filename)


if __name__ == '__main__':
    main()
